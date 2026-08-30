"""Writes the monthly P/L breakdown as a printable XLSX, three tables on one sheet:
  1. Monthly summary at the top - metrics as rows, months as columns, so you can scan
     one risk metric across the whole period at a glance.
  2. By-asset detail below - months x metric on the left (3 rows per month: Trades,
     P/L, Avg P/L), symbols as columns, so performance is directly comparable across
     assets for a given month.
  3. Cumulative P/L by asset - the opposite orientation from (2): symbols as rows,
     months as columns, each cell a running total through that month, so a symbol's
     equity curve reads left-to-right in one row.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from .monthly import AssetMonthStats, MonthSummary, cumulative_pnl_by_asset

_BOLD = Font(bold=True)

_SUMMARY_METRICS = [
    ("Trades", lambda s: s.trades),
    ("Win rate", lambda s: f"{s.win_rate:.1%}"),
    ("Total P/L", lambda s: round(s.total_pnl, 2)),
    ("Avg P/L/trade", lambda s: round(s.avg_pnl, 2)),
    ("P/L Std Dev", lambda s: round(s.pnl_stdev, 2) if s.pnl_stdev is not None else "n/a"),
    ("Best trade", lambda s: round(s.best_trade, 2)),
    ("Worst trade", lambda s: round(s.worst_trade, 2)),
    ("Max drawdown", lambda s: round(s.max_drawdown, 2)),
    ("Avg risk at entry", lambda s: round(s.avg_risk_money, 2) if s.avg_risk_money is not None else "n/a"),
]

_ASSET_METRICS = [
    ("Trades", lambda a: a.trades),
    ("P/L ($)", lambda a: round(a.pnl, 2)),
    ("Avg P/L ($)", lambda a: round(a.avg_pnl, 2)),
]


def write_monthly_workbook(
    asset_stats: list[AssetMonthStats],
    summaries: list[MonthSummary],
    confidence: float,
    output_path: str | Path,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    symbols = sorted({a.symbol for a in asset_stats})
    months = [s.month for s in summaries]

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Report"

    row = _write_summary_table(ws, 1, summaries, confidence)
    row += 2  # blank separator before the by-asset table
    row = _write_by_asset_table(ws, row, asset_stats, summaries, symbols)
    row += 2  # blank separator before the cumulative-by-asset table
    _write_cumulative_by_asset_table(ws, row, asset_stats, months)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    for i in range(max(len(symbols), len(months))):
        ws.column_dimensions[chr(ord("C") + i)].width = 12

    wb.save(output_path)
    return output_path


def _write_summary_table(ws, row: int, summaries: list[MonthSummary], confidence: float) -> int:
    ws.cell(row=row, column=1, value="Monthly summary").font = _BOLD
    for col, summary in enumerate(summaries, start=2):
        ws.cell(row=row, column=col, value=summary.month).font = _BOLD
    row += 1

    for label, getter in _SUMMARY_METRICS:
        ws.cell(row=row, column=1, value=label)
        for col, summary in enumerate(summaries, start=2):
            ws.cell(row=row, column=col, value=getter(summary))
        row += 1

    var_label = f"VaR ({confidence:.0%})"
    ws.cell(row=row, column=1, value=var_label)
    for col, summary in enumerate(summaries, start=2):
        value = round(summary.var_month, 2) if summary.var_month is not None else "n/a"
        ws.cell(row=row, column=col, value=value)
    row += 1

    return row


def _write_by_asset_table(ws, row: int, asset_stats: list[AssetMonthStats], summaries: list[MonthSummary], symbols: list[str]) -> int:
    by_key = {(a.month, a.symbol): a for a in asset_stats}

    ws.cell(row=row, column=2, value="Symbol").font = _BOLD
    for col, symbol in enumerate(symbols, start=3):
        ws.cell(row=row, column=col, value=symbol).font = _BOLD
    row += 1

    for summary in summaries:
        for label, getter in _ASSET_METRICS:
            ws.cell(row=row, column=1, value=summary.month)
            ws.cell(row=row, column=2, value=label)
            for col, symbol in enumerate(symbols, start=3):
                asset = by_key.get((summary.month, symbol))
                ws.cell(row=row, column=col, value=getter(asset) if asset is not None else None)
            row += 1
        # row += 1  # blank separator before the next month's block

    return row


def _write_cumulative_by_asset_table(ws, row: int, asset_stats: list[AssetMonthStats], months: list[str]) -> int:
    cumulative = cumulative_pnl_by_asset(asset_stats, months)
    symbols = sorted(cumulative.keys())

    ws.cell(row=row, column=1, value="Cumulative P/L by asset").font = _BOLD
    for col, month in enumerate(months, start=2):
        ws.cell(row=row, column=col, value=month).font = _BOLD
    row += 1

    for symbol in symbols:
        ws.cell(row=row, column=1, value=symbol)
        for col, month in enumerate(months, start=2):
            ws.cell(row=row, column=col, value=round(cumulative[symbol][month], 2))
        row += 1

    return row
