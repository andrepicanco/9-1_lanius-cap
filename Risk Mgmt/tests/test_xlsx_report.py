from pathlib import Path

from openpyxl import load_workbook

from risk_mgmt.monthly import AssetMonthStats, MonthSummary
from risk_mgmt.xlsx_report import write_monthly_workbook


def _sample_data():
    asset_stats = [
        AssetMonthStats(month="2026-01", symbol="DE40", trades=1, pnl=20.0, avg_pnl=20.0),
        AssetMonthStats(month="2026-01", symbol="US500", trades=2, pnl=60.0, avg_pnl=30.0),
        AssetMonthStats(month="2026-02", symbol="US500", trades=1, pnl=30.0, avg_pnl=30.0),
    ]
    summaries = [
        MonthSummary(month="2026-01", trades=3, win_rate=2 / 3, total_pnl=80.0, avg_pnl=80.0 / 3,
                     pnl_stdev=74.86, best_trade=100.0, worst_trade=-40.0, max_drawdown=40.0,
                     avg_risk_money=50.0, var_month=55.2),
        MonthSummary(month="2026-02", trades=1, win_rate=1.0, total_pnl=30.0, avg_pnl=30.0,
                     pnl_stdev=None, best_trade=30.0, worst_trade=30.0, max_drawdown=0.0,
                     avg_risk_money=None, var_month=None),
    ]
    return asset_stats, summaries


def _cell_values(ws, max_row: int, max_col: int) -> list[list]:
    return [[ws.cell(row=r, column=c).value for c in range(1, max_col + 1)] for r in range(1, max_row + 1)]


def _write(tmp_path: Path):
    asset_stats, summaries = _sample_data()
    output_path = tmp_path / "monthly_report.xlsx"
    write_monthly_workbook(asset_stats, summaries, confidence=0.95, output_path=output_path)
    return output_path


def test_summary_table_has_months_as_columns_and_metrics_as_rows(tmp_path: Path):
    output_path = _write(tmp_path)

    wb = load_workbook(output_path)
    ws = wb["Monthly Report"]
    rows = _cell_values(ws, 12, 3)

    assert rows[0][:3] == ["Monthly summary", "2026-01", "2026-02"]
    labels = [row[0] for row in rows[:11]]
    for expected in ["Trades", "Win rate", "Total P/L", "Avg P/L/trade", "P/L Std Dev",
                      "Best trade", "Worst trade", "Max drawdown", "Avg risk at entry", "VaR (95%)"]:
        assert expected in labels


def test_summary_table_values_align_with_the_right_month_column(tmp_path: Path):
    output_path = _write(tmp_path)

    wb = load_workbook(output_path)
    ws = wb["Monthly Report"]
    rows = _cell_values(ws, 12, 3)
    by_label = {row[0]: row for row in rows}

    assert by_label["Trades"][1:3] == [3, 1]
    assert by_label["Total P/L"][1:3] == [80.0, 30.0]
    assert by_label["P/L Std Dev"][1:3] == [74.86, "n/a"]  # February has a single trade
    assert by_label["Avg risk at entry"][1:3] == [50.0, "n/a"]
    assert by_label["VaR (95%)"][1:3] == [55.2, "n/a"]  # February has no VaR (single trade)


def test_by_asset_table_has_symbols_as_columns(tmp_path: Path):
    output_path = _write(tmp_path)

    wb = load_workbook(output_path)
    ws = wb["Monthly Report"]
    all_rows = _cell_values(ws, ws.max_row, 4)

    header_row = next(row for row in all_rows if row[1] == "Symbol")
    assert header_row[2:4] == ["DE40", "US500"]  # sorted alphabetically


def test_by_asset_table_rows_grouped_by_month_and_metric(tmp_path: Path):
    output_path = _write(tmp_path)

    wb = load_workbook(output_path)
    ws = wb["Monthly Report"]
    all_rows = _cell_values(ws, ws.max_row, 4)

    jan_pnl_row = next(row for row in all_rows if row[0] == "2026-01" and row[1] == "P/L ($)")
    assert jan_pnl_row[2:4] == [20.0, 60.0]  # DE40, US500

    jan_trades_row = next(row for row in all_rows if row[0] == "2026-01" and row[1] == "Trades")
    assert jan_trades_row[2:4] == [1, 2]


def test_missing_symbol_for_a_month_is_blank_not_zero(tmp_path: Path):
    output_path = _write(tmp_path)

    wb = load_workbook(output_path)
    ws = wb["Monthly Report"]
    all_rows = _cell_values(ws, ws.max_row, 4)

    # DE40 never traded in February - its column should be blank on the Feb rows, not 0
    feb_pnl_row = next(row for row in all_rows if row[0] == "2026-02" and row[1] == "P/L ($)")
    assert feb_pnl_row[2] is None  # DE40 column
    assert feb_pnl_row[3] == 30.0  # US500 column


def test_cumulative_by_asset_table_has_symbols_as_rows_and_months_as_columns(tmp_path: Path):
    output_path = _write(tmp_path)

    wb = load_workbook(output_path)
    ws = wb["Monthly Report"]
    all_rows = _cell_values(ws, ws.max_row, 3)

    header_row = next(row for row in all_rows if row[0] == "Cumulative P/L by asset")
    assert header_row[1:3] == ["2026-01", "2026-02"]

    # DE40 only traded in January (pnl 20.0) - February carries that total forward.
    de40_row = next(row for row in all_rows if row[0] == "DE40")
    assert de40_row[1:3] == [20.0, 20.0]

    # US500: Jan 60.0 total, then +30.0 in Feb -> running total 90.0
    us500_row = next(row for row in all_rows if row[0] == "US500")
    assert us500_row[1:3] == [60.0, 90.0]


def test_creates_output_directory_if_missing(tmp_path: Path):
    nested = tmp_path / "does" / "not" / "exist" / "report.xlsx"
    asset_stats, summaries = _sample_data()

    write_monthly_workbook(asset_stats, summaries, confidence=0.95, output_path=nested)

    assert nested.exists()
